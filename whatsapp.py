import pyautogui
import pygetwindow
import openpyxl
import time

# Find WhatsApp Desktop window
whatsapp_window = pygetwindow.getWindowsWithTitle('WhatsApp')[0]  # Assumes the window title is 'WhatsApp'
whatsapp_window.activate()

# Wait for WhatsApp window to be active
time.sleep(2)

# Load the Excel file
excel_file_path = 'C:\\Users\\LENOVO\\Downloads\\Book1.xlsm'  # Replace with the path to your Excel file
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Get the last row with data in column A
last_row = sheet.max_row

# Loop through each row in Excel
for row in range(1, last_row + 1):
    # Get phone number from column A and message from column B
    phone_number = str(sheet.cell(row=row, column=1).value)
    message = str(sheet.cell(row=row, column=2).value)
    
    # Click on the input field to focus
    input_field_position = (500, 800)  # Adjust these coordinates based on your screen
    pyautogui.click(input_field_position)
    
    # Typing the phone number and message
    pyautogui.typewrite(f"{phone_number} {message}")
    
    # Simulate pressing 'Enter' to send the message
    pyautogui.press('enter')
    
    # Wait for a few seconds before sending the next message (adjust as needed)
    time.sleep(2)

# Close the Excel file
wb.close()
